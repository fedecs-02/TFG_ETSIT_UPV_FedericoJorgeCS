
import os
import re
import time
import json
import joblib
import requests
import numpy as np
import pandas as pd
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# 1. Configuración
BASE_FEB            = "https://baloncestoenvivo.feb.es"
DIR_DATA            = "feb_data"
DIR_MODELOS         = os.path.join(DIR_DATA, "modelos_entrenados")
CHROMEDRIVER_PATH   = r"C:\Users\feder\Downloads\CHROMEDRIVER\chromedriver.exe"
SELENIUM_TIMEOUT    = 12
HEADLESS            = True
FILE_HISTORICO      = os.path.join(DIR_DATA, "estadisticas_jugadoras_vf.csv")
FILE_MATCHES_2025   = os.path.join(DIR_DATA, "matches_2025.json")
FILE_PLAYERS_2025   = os.path.join(DIR_DATA, "players_stats_2025.csv")
FILE_EQUIPOS_2025   = os.path.join(DIR_DATA, "stats_equipos_2025.csv")
FILE_EXCEL_OUT      = os.path.join(DIR_DATA, "validacion_2025_escenarios.xlsx")
LEAGUE_ID   = 4
SEASON      = 2025
LEAGUE_NM   = "lfendesa"
PESOS_PTC = {
    'puntos': 1.0, 'tap_fav': 0.91, 'reb_def': 0.58, 'reb_of': 0.92,
    'recup': 0.86, 'asist': 0.48, 'faltas_r': 0.23, 'tc_fail': -0.91,
    'tl_fail': -0.57, 'perd': -0.86, 'faltas_c': -0.23
}
TARGETS = [
    'puntos', 'asist', 'reb_of', 'reb_def', 'recup', 'tap_fav',
    'perd', 'faltas_c', 'faltas_r', 't2_int', 't2_met',
    't3_int', 't3_met', 'tl_int', 'tl_met', 'eFG', 'USG'
]
VARS_A_COMPARAR = [
    'puntos', 'asist', 'reb_of', 'reb_def', 'recup',
    'tap_fav', 'perd', 'faltas_c', 'faltas_r', 't2_int', 't2_met',
    't3_int', 't3_met',
    'tl_int', 'tl_met', 'eFG', 'USG'
]
URL_STATS_EQUIPOS = "https://baloncestoenvivo.feb.es/estadisticas/lfendesa/4/2025"
FASE_ID = "88870" 
os.makedirs(DIR_DATA, exist_ok=True)

#2. Extracción de partidos de la Temporada 2025
def _extract_matches_from_html(html: str):
    soup = BeautifulSoup(html, "html.parser")
    matches = []
    for a in soup.find_all("a", href=True):
        href = a["href"]
        link = href if href.startswith("http") else BASE_FEB + href if href.startswith("/") else BASE_FEB + "/" + href
        m = re.search(r'[Pp]artido\.aspx\?p=(\d+)', href)
        if m:
            pid = int(m.group(1))
            local = visitante = marcador = None
            tr = a.find_parent("tr")
            if tr:
                tds = tr.find_all("td")
                if len(tds) >= 3:
                    local     = tds[0].get_text(strip=True)
                    marcador  = tds[1].get_text(strip=True)
                    visitante = tds[2].get_text(strip=True)
            matches.append({
                "id_partido": pid, "local": local,
                "marcador": marcador, "visitante": visitante, "link": link
            })
    unique = {it["id_partido"]: it for it in matches if it["id_partido"]}
    return list(unique.values())

def _selenium_extract_matches(friendly_url: str):
    opts = Options()
    if HEADLESS:
        opts.add_argument("--headless=new")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-gpu")
    svc = Service(CHROMEDRIVER_PATH) if os.path.exists(CHROMEDRIVER_PATH) else Service()
    driver = webdriver.Chrome(service=svc, options=opts)
    try:
        driver.get(friendly_url)
        WebDriverWait(driver, SELENIUM_TIMEOUT).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "a[href*='Partido.aspx']"))
        )
        matches = _extract_matches_from_html(driver.page_source)
    finally:
        driver.quit()
    unique = {it["id_partido"]: it for it in matches if it["id_partido"]}
    return list(unique.values())


def extraer_partidos_2025():
    params_url = f"{BASE_FEB}/calendario.aspx"
    params = {"g": LEAGUE_ID, "t": SEASON, "nm": LEAGUE_NM}
    try:
        r = requests.get(params_url, params=params, timeout=15)
        r.raise_for_status()
        html = r.text
    except Exception as e:
        print(f" Error. requests falló ({e}).")
        html = None
    matches = _extract_matches_from_html(html) if html else []
    if not matches:
        friendly = f"{BASE_FEB}/calendario/{LEAGUE_NM}/{LEAGUE_ID}/{SEASON}"
        print(f" Sin partidos via requests. Se va a usar Selenium en {friendly}")
        try:
            matches = _selenium_extract_matches(friendly)
        except Exception as e:
            print(f"Error. Selenium falló: {e}")
            matches = []
    for m in matches:
        m["liga_clave"]  = "LF_ENDESA"
        m["liga_id"]     = LEAGUE_ID
        m["temporada"]   = SEASON
    matches = sorted(matches, key=lambda x: x["id_partido"] or 0)
    print(f"  → {len(matches)} partidos encontrados.")
    with open(FILE_MATCHES_2025, "w", encoding="utf-8") as f:
        json.dump(matches, f, ensure_ascii=False, indent=2)
    return matches

# 3. Scraping de las estadísticas de las jugadoras
def _init_driver():
    opts = Options()
    opts.add_argument("--headless=new")
    opts.add_argument("--window-size=1920,1080")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--no-sandbox")
    opts.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    )
    svc = Service(ChromeDriverManager().install()) 
    return webdriver.Chrome(service=svc, options=opts)
def _split_shoot(text):
    clean = text.strip().split(' ')[0]
    if '/' in clean:
        parts = clean.split('/')
        try:
            return int(parts[0]), int(parts[1])
        except:
            return 0, 0
    return 0, 0
def _parse_players_from_table(soup_table, match_id, team_name, is_local):
    players = []
    tbody = soup_table.find("tbody")
    if not tbody:
        return []
    for row in tbody.find_all("tr"):
        if "row-total" in row.get("class", []) or not row.find("td"):
            continue
        cols = row.find_all("td")
        if len(cols) < 15:
            continue
        try:
            es_titular  = 1 if "*" in cols[0].text else 0
            dorsal      = cols[1].text.strip()
            nombre      = cols[2].text.strip()
            if not nombre:
                continue
            minutos     = cols[3].text.strip()
            pts         = int(cols[4].text.strip() or 0)
            t2_m, t2_i = _split_shoot(cols[5].text)
            t3_m, t3_i = _split_shoot(cols[6].text)
            tc_m, tc_i = _split_shoot(cols[7].text)
            tl_m, tl_i = _split_shoot(cols[8].text)
            reb_of      = int(cols[9].text.strip()  or 0)
            reb_def     = int(cols[10].text.strip() or 0)
            reb_tot     = int(cols[11].text.strip() or 0)
            asist       = int(cols[12].text.strip() or 0)
            recup       = int(cols[13].text.strip() or 0)
            perd        = int(cols[14].text.strip() or 0)
            tap_fav     = int(cols[15].text.strip() or 0) if len(cols) > 15 else 0
            tap_con     = int(cols[16].text.strip() or 0) if len(cols) > 16 else 0
            mates       = int(cols[17].text.strip() or 0) if len(cols) > 17 else 0
            faltas_c    = int(cols[18].text.strip() or 0) if len(cols) > 18 else 0
            faltas_r    = int(cols[19].text.strip() or 0) if len(cols) > 19 else 0
            val         = int(cols[20].text.strip() or 0) if len(cols) > 20 else 0
            plus_minus  = int(cols[21].text.strip() or 0) if len(cols) > 21 else 0

            players.append({
                "id_partido": match_id,
                "equipo": team_name,
                "es_local": "Local" if is_local else "Visitante",
                "nombre": nombre, "dorsal": dorsal, "titular": es_titular,
                "minutos": minutos, "puntos": pts,
                "t2_met": t2_m, "t2_int": t2_i,
                "t3_met": t3_m, "t3_int": t3_i,
                "tl_met": tl_m, "tl_int": tl_i,
                "reb_of": reb_of, "reb_def": reb_def, "reb_tot": reb_tot,
                "asist": asist, "recup": recup, "perd": perd,
                "tap_fav": tap_fav, "tap_con": tap_con, "mates": mates,
                "faltas_c": faltas_c, "faltas_r": faltas_r,
                "val": val, "mas_menos": plus_minus,
                "liga": "LF_ENDESA", "temporada": SEASON
            })
        except Exception:
            continue
    return players
def _process_match(driver, match_info):
    url      = match_info.get("link") or f"{BASE_FEB}/Partido.aspx?p={match_info['id_partido']}"
    match_id = match_info["id_partido"]
    nombre_local     = match_info.get("local", "Local")
    nombre_visitante = match_info.get("visitante", "Visitante")

    try:
        driver.get(url)
        WebDriverWait(driver, 8).until(
            EC.presence_of_element_located((By.CLASS_NAME, "responsive-scroll"))
        )
        soup       = BeautifulSoup(driver.page_source, "html.parser")
        containers = soup.find_all("div", class_="responsive-scroll")
        valid_tables = [
            c.find("table") for c in containers
            if c.find("table") and c.find("table").find("th", class_="dorsal")
        ]
        if len(valid_tables) >= 2:
            jugadoras = (
                _parse_players_from_table(valid_tables[0], match_id, nombre_local, True) +
                _parse_players_from_table(valid_tables[1], match_id, nombre_visitante, False)
            )
            print(f"   [OK] ID {match_id}: {len(jugadoras)} jugadoras.")
            return jugadoras
        else:
            print(f"   [X]  ID {match_id}: Sin tablas válidas.")
            return []
    except Exception as e:
        print(f"   [ERR] ID {match_id}: {e}")
        return []


def extraer_jugadoras_2025(matches):
    # Si ya existe, cargamos y añadimos solo los nuevos
    partidos_ya_procesados = set()
    all_data = []
    if os.path.exists(FILE_PLAYERS_2025):
        df_exist = pd.read_csv(FILE_PLAYERS_2025)
        partidos_ya_procesados = set(df_exist["id_partido"].unique())
        all_data = df_exist.to_dict("records")
        print(f"  [INFO] {len(partidos_ya_procesados)} partidos ya procesados. Solo scraperemos los nuevos.")

    matches_pendientes = [m for m in matches if m["id_partido"] not in partidos_ya_procesados]
    print(f"  [INFO] {len(matches_pendientes)} partidos pendientes.")

    if not matches_pendientes:
        print("  → Todo ya procesado.")
        return pd.read_csv(FILE_PLAYERS_2025)

    driver = _init_driver()
    try:
        for m in matches_pendientes:
            jugadoras = _process_match(driver, m)
            if jugadoras:
                all_data.extend(jugadoras)
                pd.DataFrame(all_data).to_csv(FILE_PLAYERS_2025, index=False, encoding="utf-8-sig")
            time.sleep(0.3)
    finally:
        driver.quit()

    df = pd.DataFrame(all_data)
    df.to_csv(FILE_PLAYERS_2025, index=False, encoding="utf-8-sig")
    print(f"  → {len(df)} registros guardados en {FILE_PLAYERS_2025}")
    return df

# 4. LIMPIEZA DE DATOS

def _time_to_float(time_str):
    if not isinstance(time_str, str):
        return 0.0
    time_str = str(time_str).strip()
    if time_str in ("00:00", "0", "", "nan"):
        return 0.0
    if ":" in time_str:
        try:
            parts = time_str.split(":")
            return int(parts[0]) + int(parts[1]) / 60
        except:
            return 0.0
    try:
        return float(time_str)
    except:
        return 0.0


def limpiar_y_preparar(df):
    df["minutos"] = df["minutos"].astype(str).str.strip()
    df_clean = df[
        (df["minutos"] != "00:00") &
        (df["minutos"] != "0") &
        (df["minutos"] != "") &
        (df["minutos"] != "nan")
    ].copy()
    df_clean["minutos_float"] = df_clean["minutos"].apply(_time_to_float)
    df_clean = df_clean[df_clean["minutos_float"] > 0].copy()
    df_clean["nombre"] = df_clean["nombre"].str.upper().str.strip()
    print(f"  → {len(df_clean)} registros válidos (de {len(df)} totales).")
    return df_clean

# 5. ESTADÍSTICAS DE EQUIPOS 2025

def obtener_stats_equipos_2025():
    headers = {
        "User-Agent":
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    }
    session = requests.Session()

    # GET inicial para obtener tokens del formulario ASP.NET
    resp = session.get(URL_STATS_EQUIPOS, headers=headers, timeout=20)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.content, "html.parser")

    def _token(name):
        tag = soup.find("input", {"id": name})
        return tag["value"] if tag else ""

    payload = {
        "__EVENTTARGET":    "_ctl0:MainContentPlaceHolderMaster:fasesGruposDropDownList",
        "__EVENTARGUMENT":  "",
        "__VIEWSTATE":      _token("__VIEWSTATE"),
        "__VIEWSTATEGENERATOR": _token("__VIEWSTATEGENERATOR"),
        "__EVENTVALIDATION":    _token("__EVENTVALIDATION"),
        "_ctl0:MainContentPlaceHolderMaster:fasesGruposDropDownList": FASE_ID
    }
    resp_post = session.post(URL_STATS_EQUIPOS, headers=headers, data=payload, timeout=20)
    resp_post.raise_for_status()
    soup2 = BeautifulSoup(resp_post.content, "html.parser")

    datos = []
    for fila in soup2.find_all("tr"):
        td_nombre = fila.find("td", class_="nombre equipo")
        if not td_nombre:
            continue
        nombre   = td_nombre.text.strip()
        td_pj    = fila.find("td", class_="partidos")
        partidos = int(td_pj.text.strip()) if td_pj and td_pj.text.strip().isdigit() else 1

        def get_tiros(clase):
            try:
                tot = fila.find("td", class_=clase).find("span", class_="tot").text
                met, intt = map(int, tot.split("/"))
                return met / partidos, intt / partidos
            except:
                return 0.0, 0.0

        def get_med(clase):
            try:
                val = fila.find("td", class_=clase).find("span", class_="med").text
                return float(val.replace(",", "."))
            except:
                return 0.0

        t2_m, t2_i  = get_tiros("tiros dos")
        t3_m, t3_i  = get_tiros("tiros tres")
        tl_m, tl_i  = get_tiros("tiros libres")
        reb_of       = get_med("rebotes ofensivos")
        reb_def      = get_med("rebotes defensivos")
        perd         = get_med("perdidas")

        fga  = t2_i + t3_i
        pace = t2_i + t3_i + perd + (0.44 * tl_i) - reb_of
        efg  = 100 * (t2_m + 1.5 * t3_m) / fga if fga > 0 else 0
        tov  = 100 * (perd / pace)              if pace > 0 else 0
        ftr  = 100 * tl_m / fga                if fga  > 0 else 0

        datos.append({
            "equipo":   nombre,
            "PACE":     round(pace, 2),
            "eFG_pct":  round(efg, 2),
            "TOV_pct":  round(tov, 2),
            "FTR":      round(ftr, 2),
            "ORB_med":  round(reb_of,  2),
            "DRB_med":  round(reb_def, 2),
        })

    df_eq = pd.DataFrame(datos)
    print(f"  → {len(df_eq)} equipos obtenidos.")
    return df_eq
# 6. CLUSTERIZACIÓN DE EQUIPOS (= SIMULADOR_V3.py → predecir_cluster)

def clusterizar_equipos(df_stats, kmeans, scaler):
    # Media global de ORB/DRB para usar como rival promedio
    mean_orb = df_stats["ORB_med"].mean()
    mean_drb = df_stats["DRB_med"].mean()

    resultados = []
    for _, row in df_stats.iterrows():
        equipo = row["equipo"]
        # ORB% y DRB% relativos (igual que SIMULADOR_V4 con rival = media global)
        orb_pct = round(100 * row["ORB_med"] / (row["ORB_med"] + mean_drb), 1) if (row["ORB_med"] + mean_drb) > 0 else 50.0
        drb_pct = round(100 * row["DRB_med"] / (row["DRB_med"] + mean_orb), 1) if (row["DRB_med"] + mean_orb) > 0 else 50.0

        datos_modelo = {
            "EFG":    row["eFG_pct"],
            "TOV":    row["TOV_pct"],
            "ORB":    orb_pct,
            "DRB":    drb_pct,
            "FT/FGA": row["FTR"],
            "PACE":   row["PACE"],
        }
        df_tmp = pd.DataFrame([datos_modelo])
        try:
            orden = scaler.feature_names_in_
            df_tmp = df_tmp[orden]
        except AttributeError:
            pass  # Versión de scikit sin feature_names_in_

        try:
            datos_sc    = scaler.transform(df_tmp)
            cluster     = int(kmeans.predict(datos_sc)[0])
            distancias  = kmeans.transform(datos_sc)
            dist_centro = float(distancias[0, cluster])
        except Exception as e:
            print(f"  [WARN] Cluster fallido para {equipo}: {e} → asignando cluster 0")
            cluster, dist_centro = 0, 0.0

        resultados.append({
            "equipo":           equipo,
            "cluster":          cluster,
            "dist_centroide":   round(dist_centro, 4),
            "PACE":             row["PACE"],
        })

    df_cluster = pd.DataFrame(resultados)
    df_cluster.to_csv(FILE_EQUIPOS_2025, index=False, encoding="utf-8-sig")
    print(f"  → Clusters calculados y guardados en {FILE_EQUIPOS_2025}")
    print(df_cluster[["equipo", "cluster", "dist_centroide", "PACE"]].to_string(index=False))
    return df_cluster
    
# 7. INGENIERÍA DE FEATURES JUGADORAS
def calcular_ptc_row(row):
    tc_f = (row["t2_int"] - row["t2_met"]) + (row["t3_int"] - row["t3_met"])
    tl_f = row["tl_int"] - row["tl_met"]
    return (
        row["puntos"]   * PESOS_PTC["puntos"]   + row["tap_fav"] * PESOS_PTC["tap_fav"] +
        row["reb_def"]  * PESOS_PTC["reb_def"]  + row["reb_of"]  * PESOS_PTC["reb_of"] +
        row["recup"]    * PESOS_PTC["recup"]    + row["asist"]   * PESOS_PTC["asist"] +
        row["faltas_r"] * PESOS_PTC["faltas_r"] + tc_f * PESOS_PTC["tc_fail"] +
        tl_f * PESOS_PTC["tl_fail"] + row["perd"] * PESOS_PTC["perd"] +
        row["faltas_c"] * PESOS_PTC["faltas_c"]
    )


def generar_features_jugadoras(df_2025, df_historico, df_cluster_equipos, features_list):
    # 1. Concatenar histórico + 2025 en orden cronológico
    df_historico["temporada"] = df_historico["temporada"].astype(int)
    df_2025["temporada"]      = df_2025["temporada"].astype(int)

    # Asegurar columnas numéricas comunes
    cols_num = [
        "puntos", "t2_met", "t2_int", "t3_met", "t3_int",
        "tl_met", "tl_int", "reb_of", "reb_def", "reb_tot",
        "asist", "recup", "perd", "tap_fav", "faltas_c", "faltas_r", "minutos_float"
    ]
    df_historico = df_historico.fillna(0)
    df_2025      = df_2025.fillna(0)

    df_completo = pd.concat([df_historico, df_2025], ignore_index=True, sort=False)
    df_completo = df_completo.sort_values(["nombre", "temporada", "id_partido"]).reset_index(drop=True)

    # 2. Identificar equipo rival en cada partido
    partidos = df_completo.groupby("id_partido")["equipo"].unique()

    def get_rival(pid, my_team):
        teams = partidos.get(pid, [])
        if len(teams) == 2:
            return teams[0] if teams[1] == my_team else teams[1]
        return None

    df_completo["rival"] = df_completo.apply(lambda x: get_rival(x["id_partido"], x["equipo"]), axis=1)
    df_completo = df_completo.dropna(subset=["rival"])

    # 3. Asignar cluster y PACE del rival (de df_cluster_equipos)
    # Normalizamos nombres para el cruce
    df_cluster_equipos["equipo_norm"] = df_cluster_equipos["equipo"].str.upper().str.strip()
    df_completo["rival_norm"]         = df_completo["rival"].str.upper().str.strip()

    cluster_dict       = df_cluster_equipos.set_index("equipo_norm")["cluster"].to_dict()
    dist_centroide_dict = df_cluster_equipos.set_index("equipo_norm")["dist_centroide"].to_dict()
    pace_dict          = df_cluster_equipos.set_index("equipo_norm")["PACE"].to_dict()

    # Media global como fallback
    cluster_default       = df_cluster_equipos["cluster"].mode()[0] if len(df_cluster_equipos) > 0 else 0
    dist_centroide_default = df_cluster_equipos["dist_centroide"].mean() if len(df_cluster_equipos) > 0 else 0.0
    pace_default          = df_cluster_equipos["PACE"].mean() if len(df_cluster_equipos) > 0 else 70.0

    df_completo["cluster_def_rival"]    = df_completo["rival_norm"].map(cluster_dict).fillna(cluster_default).astype(int)
    df_completo["dist_centroide_rival"] = df_completo["rival_norm"].map(dist_centroide_dict).fillna(dist_centroide_default)
    df_completo["pace_rival"]           = df_completo["rival_norm"].map(pace_dict).fillna(pace_default)

    # 4. Calcular PTC, eFG, USG para todas las filas
    df_completo["PTC"] = df_completo.apply(calcular_ptc_row, axis=1)
    posesiones = (df_completo["minutos_float"].replace(0, 0.1) / 40) * df_completo["pace_rival"]
    df_completo["PTC_mp"] = (df_completo["PTC"] / posesiones) * 100

    uso_jugadora = df_completo["t2_int"] + df_completo["t3_int"] + 0.44 * df_completo["tl_int"] + df_completo["perd"]
    uso_jugadora = uso_jugadora.replace(0, 0.1)
    df_completo["PTS_mpu"] = (df_completo["puntos"] / (df_completo["minutos_float"].replace(0, 0.1) * uso_jugadora)) * 40

    tiros_totales = df_completo["t2_int"] + df_completo["t3_int"]
    df_completo["eFG"] = np.where(
        tiros_totales > 0,
        ((df_completo["t2_met"] + 1.5 * df_completo["t3_met"]) / tiros_totales) * 100,
        0
    )
    df_completo["USG"] = (uso_jugadora / posesiones) * 100

    # 5. Calcular L6_* y Sn_* usando shift para no incluir el partido actual
    cols_stats = [
        "puntos", "reb_tot", "asist", "PTC_mp", "minutos_float", "PTS_mpu", "eFG", "USG",
        "recup", "tap_fav", "perd", "reb_of", "reb_def", "faltas_c", "faltas_r",
        "t2_int", "t2_met", "t3_int", "t3_met", "tl_int", "tl_met"
    ]
    for col in cols_stats:
        df_completo[f"L6_{col}"] = df_completo.groupby("nombre")[col].transform(
            lambda x: x.shift().ewm(span=6, min_periods=1).mean()
        )
        df_completo[f"Sn_{col}"] = df_completo.groupby(["nombre", "temporada"])[col].transform(
            lambda x: x.shift().expanding().mean()
        )
        if col in ['PTC_mp', 'puntos']:
            df_completo[f'L6_{col}_std'] = df_completo.groupby('nombre')[col].transform(lambda x: x.shift().rolling(6, min_periods=2).std().fillna(0))
            df_completo[f'L6_{col}_min'] = df_completo.groupby('nombre')[col].transform(lambda x: x.shift().rolling(6, min_periods=1).min().fillna(0))
            if col == 'PTC_mp':
                df_completo['es_partido_malo'] = (df_completo['PTC_mp'] < 0).astype(int)
                df_completo['Sn_ratio_malos'] = df_completo.groupby(['nombre', 'temporada'])['es_partido_malo'].transform(lambda x: x.shift().expanding().mean().fillna(0))
    # 6. Filtrar: solo filas de 2025 con datos históricos presentes
    df_train = df_completo[
        (df_completo["temporada"] == SEASON) &
        df_completo["L6_puntos"].notna() &
        df_completo["Sn_puntos"].notna()
    ].copy()

    print(f"  → {len(df_train)} filas de 2025 con features completas.")

    # 7. Factor cancha
    if df_train["es_local"].dtype == object:
        df_train["es_local"] = np.where(df_train["es_local"].str.lower().str.contains("local"), 1, 0)

    # 8. One-Hot Encoding del cluster rival
    df_train["cluster_def_rival"] = df_train["cluster_def_rival"].astype(int)
    df_train = pd.get_dummies(df_train, columns=["cluster_def_rival"], drop_first=False)
    for i in range(6):
        col_name = f"cluster_def_rival_{i}"
        if col_name not in df_train.columns:
            df_train[col_name] = 0.0

    # 9. Asegurar todas las features
    for col in features_list:
        if col not in df_train.columns:
            df_train[col] = 0.0

    return df_train

# 8. PREDICCIÓN Y SIMULACIÓN DE 4 ESCENARIOS

def _optimizar_minutos(df_partido, col_eficiencia):
    mins_seguros = df_partido["minutos_float"].replace(0, 0.1)
    eficiencia   = df_partido[col_eficiencia] / mins_seguros
    pesos        = eficiencia.clip(lower=0).values

    if np.sum(pesos) == 0:
        return df_partido["minutos_float"].values

    mins_asignados  = np.zeros(len(df_partido))
    mins_restantes  = 200.0
    tope            = 35.0

    while mins_restantes > 0.1 and np.sum(pesos) > 0:
        cuotas = (pesos / np.sum(pesos)) * mins_restantes
        for i in range(len(df_partido)):
            if pesos[i] > 0:
                if mins_asignados[i] + cuotas[i] > tope:
                    mins_restantes       -= (tope - mins_asignados[i])
                    mins_asignados[i]    = tope
                    pesos[i]             = 0
                else:
                    mins_asignados[i]   += cuotas[i]
                    mins_restantes      -= cuotas[i]
    return mins_asignados


def predecir_y_simular(df_train, regresores, features_list):
    X = df_train[features_list]

    # Predicciones de todas las variables target
    for t in TARGETS:
        if t in regresores:
            df_train[f"{t}_pred"] = regresores[t].predict(X).clip(min=0)
        else:
            df_train[f"{t}_pred"] = 0.0

    # PTC Proyectado por la IA
    fallos_t2 = (df_train["t2_int_pred"] - df_train["t2_met_pred"]).clip(lower=0)
    fallos_t3 = (df_train["t3_int_pred"] - df_train["t3_met_pred"]).clip(lower=0)
    fallos_tl = (df_train["tl_int_pred"] - df_train["tl_met_pred"]).clip(lower=0)

    df_train["PTC_Proy_IA"] = (
        df_train["puntos_pred"]   * PESOS_PTC["puntos"]   +
        df_train["tap_fav_pred"]  * PESOS_PTC["tap_fav"]  +
        df_train["reb_def_pred"]  * PESOS_PTC["reb_def"]  +
        df_train["reb_of_pred"]   * PESOS_PTC["reb_of"]   +
        df_train["recup_pred"]    * PESOS_PTC["recup"]    +
        df_train["asist_pred"]    * PESOS_PTC["asist"]    +
        df_train["faltas_r_pred"] * PESOS_PTC["faltas_r"] +
        (fallos_t2 + fallos_t3)   * PESOS_PTC["tc_fail"]  +
        fallos_tl                 * PESOS_PTC["tl_fail"]   +
        df_train["perd_pred"]     * PESOS_PTC["perd"]     +
        df_train["faltas_c_pred"] * PESOS_PTC["faltas_c"]
    )

    # Minutos óptimos por equipo-partido
    df_train["Min_Optimos_IA"] = 0.0
    for (partido_id, equipo), df_partido in df_train.groupby(["id_partido", "equipo"]):
        min_opt = _optimizar_minutos(df_partido, "PTC_Proy_IA")
        df_train.loc[df_partido.index, "Min_Optimos_IA"] = min_opt

    # Construcción del DataFrame de resultados
    resultados = []
    for _, row in df_train.iterrows():
        mins_reales  = row["minutos_float"]
        mins_seguros = max(mins_reales, 0.1)
        min_opt      = row["Min_Optimos_IA"]

        fila = {
            "ID_Partido":       row["id_partido"],
            "Equipo":           row["equipo"],
            "Jugadora":         row["nombre"],
            "Min_Reales":       round(mins_reales, 1),
            "Min_Optimos_IA":   round(min_opt, 1),
        }

        # PTC
        e1_ptc = row["PTC"]
        e2_ptc = row["PTC_Proy_IA"]
        e3_ptc = (row["PTC_Proy_IA"] / mins_seguros) * min_opt
        eficiencia_media_ptc = (row["Sn_PTC_mp"] / 100) * (row["pace_rival"] / 40)
        e4_ptc = eficiencia_media_ptc * mins_reales

        fila.update({
            "E1_PTC_Real":          round(e1_ptc, 2),
            "E2_PTC_IA_MinReales":  round(e2_ptc, 2),
            "E3_PTC_IA_MinOptimos": round(e3_ptc, 2),
            "E4_PTC_Media_MinReales": round(e4_ptc, 2),
        })

        # Resto de variables
        for var in VARS_A_COMPARAR:
            e1_val = row[var]
            e2_val = row[f"{var}_pred"]
            e3_val = (row[f"{var}_pred"] / mins_seguros) * min_opt

            if f"Sn_{var}" in row.index and f"Sn_minutos_float" in row.index:
                mins_hist_seguros   = max(row["Sn_minutos_float"], 0.1)
                eficiencia_media_v  = row[f"Sn_{var}"] / mins_hist_seguros
                e4_val              = eficiencia_media_v * mins_reales
            else:
                e4_val = 0.0

            fila.update({
                f"E1_{var}_Real":   round(e1_val, 2),
                f"E2_{var}_IA":     round(e2_val, 2),
                f"E3_{var}_Optimo": round(e3_val, 2),
                f"E4_{var}_Media":  round(e4_val, 2),
            })

        resultados.append(fila)

    df_esc = pd.DataFrame(resultados)
    print(f"  → {len(df_esc)} filas generadas (jugadoras × partidos).")
    return df_esc

def exportar_resultados(df_esc):

    # Estilo de cabecera
    HEADER_FILL = PatternFill("solid", fgColor="1F3864")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
    HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
    DATA_ALIGN   = Alignment(horizontal="center", vertical="center")

    def _aplicar_estilo(ws, df):
        # Cabeceras
        for cell in ws[1]:
            cell.font      = HEADER_FONT
            cell.fill      = HEADER_FILL
            cell.alignment = HEADER_ALIGN
        # Datos
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = DATA_ALIGN
        # Ancho de columnas
        for col_idx, col in enumerate(df.columns, 1):
            max_len = max(len(str(col)), df[col].astype(str).map(len).max() if len(df) > 0 else 10)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 30)
        # Fila de cabecera más alta
        ws.row_dimensions[1].height = 35

    equipos = sorted(df_esc["Equipo"].unique())

    with pd.ExcelWriter(FILE_EXCEL_OUT, engine="openpyxl") as writer:
        # Hoja global
        df_esc.to_excel(writer, sheet_name="Todos", index=False)
        _aplicar_estilo(writer.sheets["Todos"], df_esc)

        # Una hoja por equipo
        for equipo in equipos:
            df_eq = df_esc[df_esc["Equipo"] == equipo].copy()
            hoja_nombre = equipo[:31].replace("/", "-").replace("\\", "-").replace("?", "").replace("*", "")
            df_eq.to_excel(writer, sheet_name=hoja_nombre, index=False)
            _aplicar_estilo(writer.sheets[hoja_nombre], df_eq)

    print(f"  → Excel guardado: {FILE_EXCEL_OUT}")
    print(f"  → Hojas: 'Todos' + {len(equipos)} equipos.")

def main():
    print("=" * 65)
    print("  VALIDACIÓN MODELO XGBOOST — LF ENDESA 2025")
    print("=" * 65)

 
    print("\n[INFO] Cargando modelos entrenados...")
    regresores  = joblib.load(os.path.join(DIR_MODELOS, "regresores.pkl"))
    features    = joblib.load(os.path.join(DIR_MODELOS, "features.pkl"))
    kmeans      = joblib.load(os.path.join(DIR_MODELOS, "kmeans_model.pkl"))
    scaler      = joblib.load(os.path.join(DIR_MODELOS, "scaler_cluster.pkl"))
    print(f"  → Modelos cargados. Features del modelo: {len(features)}")

    if os.path.exists(FILE_MATCHES_2025):
        print(f"\n[FASE 1] Cargando partidos desde caché ({FILE_MATCHES_2025})...")
        with open(FILE_MATCHES_2025, "r", encoding="utf-8") as f:
            matches = json.load(f)
        print(f"  → {len(matches)} partidos cargados.")
    else:
        matches = extraer_partidos_2025()

    if not matches:
        print("[ERROR] No se encontraron partidos. Abortando.")
        return

    df_raw = extraer_jugadoras_2025(matches)

    if df_raw.empty:
        print("[ERROR] Sin datos de jugadoras. Abortando.")
        return

    df_2025 = limpiar_y_preparar(df_raw)

    if os.path.exists(FILE_EQUIPOS_2025):
        print(f"\n[FASE 4A] Cargando stats de equipos desde caché ({FILE_EQUIPOS_2025})...")
        df_stats_equipos = pd.read_csv(FILE_EQUIPOS_2025)
        # Si el CSV ya tiene columna cluster es que ya fue clusterizado
        if "cluster" not in df_stats_equipos.columns:
            df_stats_equipos_raw = obtener_stats_equipos_2025()
            df_cluster_equipos   = clusterizar_equipos(df_stats_equipos_raw, kmeans, scaler)
        else:
            df_cluster_equipos = df_stats_equipos[["equipo", "cluster", "dist_centroide", "PACE"]].copy()
            print(f"  → {len(df_cluster_equipos)} equipos cargados con cluster.")
    else:
        df_stats_equipos_raw = obtener_stats_equipos_2025()
        df_cluster_equipos   = clusterizar_equipos(df_stats_equipos_raw, kmeans, scaler)

    print(f"\n[INFO] Cargando histórico de jugadoras ({FILE_HISTORICO})...")
    df_historico = pd.read_csv(FILE_HISTORICO).fillna(0)
    df_historico["nombre"] = df_historico["nombre"].str.upper().str.strip()
    print(f"  → {len(df_historico)} registros históricos.")

    df_train = generar_features_jugadoras(df_2025, df_historico, df_cluster_equipos, features)

    if df_train.empty:
        print("[ERROR] Sin datos de entrenamiento para 2025. Abortando.")
        return

    df_escenarios = predecir_y_simular(df_train, regresores, features)

    exportar_resultados(df_escenarios)

    # Guardar también CSV de respaldo
    csv_out = FILE_EXCEL_OUT.replace(".xlsx", ".csv")
    df_escenarios.to_csv(csv_out, index=False, encoding="utf-8-sig")
    print(f"  → CSV de respaldo: {csv_out}")

    print("\n" + "=" * 65)
    print("  PROCESO FINALIZADO CON ÉXITO")
    print(f"  Partidos procesados : {df_escenarios['ID_Partido'].nunique()}")
    print(f"  Jugadoras únicas    : {df_escenarios['Jugadora'].nunique()}")
    print(f"  Equipos             : {df_escenarios['Equipo'].nunique()}")
    print(f"  Excel generado      : {os.path.abspath(FILE_EXCEL_OUT)}")
    print("=" * 65)


if __name__ == "__main__":
    main()
